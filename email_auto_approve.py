#!/usr/bin/env python3
"""
ServiceNow邮件自动批准程序

监控Thunderbird邮件目录下的Archive/ServiceNow/NeedApprove文件夹，
自动识别并回复批准邮件。
"""

import os
import sys
import time
import email
import json
import logging
import subprocess
import platform
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate, make_msgid
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import configparser

class EmailAutoApprover:
    def __init__(self, config_file='config.ini'):
        self.config_file = config_file
        self.config = self.load_config()
        self.setup_logging()
        
    def load_config(self):
        """加载配置文件"""
        config = configparser.ConfigParser()
        if os.path.exists(self.config_file):
            config.read(self.config_file, encoding='utf-8')
        else:
            # 创建默认配置
            config['DEFAULT'] = {
                'thunderbird_profile_path': self.get_default_thunderbird_path(),
                'watch_folder': 'Archive/ServiceNow/NeedApprove',
                'processed_emails': 'processed_emails.json',
                'log_level': 'INFO',
                'auto_approve_enabled': 'True'
            }
            config['EMAIL'] = {
                'from_name': 'Edward Li',
                'from_email': 'eli23@lululemon.com',
                'approval_message': 'Ref:MSG85395759'
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                config.write(f)
        return config
    
    def get_default_thunderbird_path(self):
        """获取默认的Thunderbird配置文件路径"""
        system = platform.system()
        user_home = Path.home()
        
        if system == "Windows":
            return str(user_home / "AppData" / "Roaming" / "Thunderbird" / "Profiles")
        elif system == "Darwin":  # macOS
            return str(user_home / "Library" / "Thunderbird" / "Profiles")
        else:  # Linux
            return str(user_home / ".thunderbird")
    
    def find_thunderbird_mail_folder(self, folder_path):
        """查找Thunderbird邮件文件夹的实际路径"""
        profile_path = self.config.get('DEFAULT', 'thunderbird_profile_path')
        
        if not os.path.exists(profile_path):
            self.logger.error(f"Thunderbird配置路径不存在: {profile_path}")
            return None
        
        # 查找所有配置文件目录
        for item in os.listdir(profile_path):
            profile_dir = os.path.join(profile_path, item)
            if not os.path.isdir(profile_dir):
                continue
                
            # 检查是否是IMAP路径 (webaccountMail/server/folder)
            if folder_path.startswith('webaccountMail/'):
                path_parts = folder_path.split('/')
                if len(path_parts) >= 3:  # webaccountMail/server/folder
                    # webaccountMail直接在profile目录下，不在Mail目录下
                    webaccount_dir = os.path.join(profile_dir, path_parts[0])  # webaccountMail
                    server_dir = os.path.join(webaccount_dir, path_parts[1])     # outlook.office365.com
                    
                    if os.path.exists(server_dir):
                        # 构建文件夹路径
                        remaining_path = '/'.join(path_parts[2:])  # ServiceNow
                        result_path = self._find_folder_recursive(server_dir, remaining_path.split('/'))
                        if result_path:
                            return result_path
            
            # 查找Local Folders和其他Mail目录下的文件夹
            else:
                # 如果路径以 Mail/ 开头，去掉这个前缀
                if folder_path.startswith('Mail/'):
                    folder_path = folder_path[5:]  # 去掉"Mail/"
                
                mail_base_dir = os.path.join(profile_dir, 'Mail')
                if not os.path.exists(mail_base_dir):
                    continue
                
                # 查找Local Folders
                if folder_path.startswith('Local Folders/'):
                    local_folders_dir = os.path.join(mail_base_dir, 'Local Folders')
                    if os.path.exists(local_folders_dir):
                        remaining_path = folder_path[14:]  # 去掉"Local Folders/"
                        
                        result_path = self._find_folder_recursive(local_folders_dir, remaining_path.split('/'))
                        if result_path:
                            return result_path
                
                # 其他路径处理
                result_path = self._find_folder_recursive(mail_base_dir, folder_path.split('/'))
                if result_path:
                    return result_path
        
        return None
    
    def _find_folder_recursive(self, current_path, folder_parts):
        """递归查找文件夹路径"""
        if not folder_parts:
            return current_path
        
        if not os.path.exists(current_path):
            return None
        
        folder_name = folder_parts[0]
        remaining_parts = folder_parts[1:]
        
        # 尝试多种查找方式
        potential_paths = []
        
        try:
            # 如果当前路径是文件，检查是否有对应的.sbd目录
            if os.path.isfile(current_path):
                sbd_path = current_path + '.sbd'
                if os.path.exists(sbd_path) and os.path.isdir(sbd_path):
                    current_path = sbd_path
                else:
                    return None
            
            if not os.path.isdir(current_path):
                return None
            
            # 1. 直接匹配目录
            direct_path = os.path.join(current_path, folder_name)
            if os.path.exists(direct_path):
                if os.path.isdir(direct_path):
                    potential_paths.append(direct_path)
                elif os.path.isfile(direct_path):
                    # 如果是最后一级并且是文件，返回包含该文件的目录
                    if not remaining_parts:
                        return current_path  # 返回包含邮件文件的目录
                    # 检查是否有对应的.sbd目录
                    sbd_path = direct_path + '.sbd'
                    if os.path.exists(sbd_path) and os.path.isdir(sbd_path):
                        potential_paths.append(sbd_path)
            
            # 2. 查找.sbd目录
            sbd_path = os.path.join(current_path, folder_name + '.sbd')
            if os.path.exists(sbd_path) and os.path.isdir(sbd_path):
                potential_paths.append(sbd_path)
            
            # 3. 忽略大小写匹配
            for item in os.listdir(current_path):
                item_path = os.path.join(current_path, item)
                
                if item.lower() == folder_name.lower():
                    if os.path.isdir(item_path):
                        potential_paths.append(item_path)
                    elif os.path.isfile(item_path):
                        if not remaining_parts:
                            return current_path  # 返回包含邮件文件的目录
                        sbd_path = item_path + '.sbd'
                        if os.path.exists(sbd_path) and os.path.isdir(sbd_path):
                            potential_paths.append(sbd_path)
                
                # 检查.sbd目录
                elif item.lower() == (folder_name.lower() + '.sbd') and os.path.isdir(item_path):
                    potential_paths.append(item_path)
            
            # 尝试每个找到的路径
            for path in potential_paths:
                result = self._find_folder_recursive(path, remaining_parts)
                if result:
                    return result
        
        except (OSError, PermissionError) as e:
            self.logger.warning(f"访问路径时出错 {current_path}: {e}")
        
        return None
    
    def setup_logging(self):
        """设置日志"""
        log_level = self.config.get('DEFAULT', 'log_level', fallback='INFO')
        logging.basicConfig(
            level=getattr(logging, log_level),
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('email_auto_approve.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def load_processed_emails(self):
        """加载已处理的邮件列表"""
        processed_file = self.config.get('DEFAULT', 'processed_emails')
        try:
            if os.path.exists(processed_file):
                with open(processed_file, 'r', encoding='utf-8') as f:
                    return set(json.load(f))
        except Exception as e:
            self.logger.error(f"加载已处理邮件列表失败: {e}")
        return set()
    
    def save_processed_emails(self, processed_emails):
        """保存已处理的邮件列表"""
        processed_file = self.config.get('DEFAULT', 'processed_emails')
        try:
            with open(processed_file, 'w', encoding='utf-8') as f:
                json.dump(list(processed_emails), f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.error(f"保存已处理邮件列表失败: {e}")
    
    def load_processing_summary(self):
        """加载处理汇总记录"""
        summary_file = 'processing_summary.json'
        try:
            if os.path.exists(summary_file):
                with open(summary_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            self.logger.error(f"加载处理汇总记录失败: {e}")
        return []
    
    def save_processing_summary(self, summary_data):
        """保存处理汇总记录"""
        summary_file = 'processing_summary.json'
        try:
            with open(summary_file, 'w', encoding='utf-8') as f:
                json.dump(summary_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.error(f"保存处理汇总记录失败: {e}")
    
    def add_to_processing_summary(self, email_info, processed_time, ticket_number):
        """添加记录到处理汇总"""
        summary_data = self.load_processing_summary()
        
        # 提取单号（RITM或CHG）
        import re
        subject = email_info.get('subject', '')
        ticket_match = re.search(r'(RITM\d+|CHG\d+)', subject, re.IGNORECASE)
        if not ticket_number and ticket_match:
            ticket_number = ticket_match.group(1)
        
        summary_record = {
            'processed_time': processed_time,
            'ticket_number': ticket_number or 'N/A',
            'short_description': email_info.get('short_description', 'N/A'),
            'subject': email_info.get('subject', 'N/A'),
            'from': email_info.get('from', 'N/A'),
            'message_id': email_info.get('message_id', 'N/A'),
            'requested_by': email_info.get('requested_by', 'N/A')
        }
        
        # 如果是China Cloud或CN-Server & DB Access Control邮件，添加额外字段
        short_desc_lower = email_info.get('short_description', '').lower()
        if (short_desc_lower.startswith('china cloud account and permission request') or 
            short_desc_lower.startswith('china cloud resource request') or
            short_desc_lower.startswith('cn-server & db access control')):
            summary_record.update({
                'environment': email_info.get('environment', 'N/A'),
                'required_permissions': email_info.get('required_permissions', 'N/A'),
                'reason_for_application': email_info.get('reason_for_application', 'N/A'),
                'is_china_cloud': True
            })
        else:
            summary_record['is_china_cloud'] = False
        
        summary_data.append(summary_record)
        self.save_processing_summary(summary_data)
        
        # 打印汇总信息
        print(f"📊 处理汇总已更新:")
        print(f"   时间: {processed_time}")
        print(f"   单号: {ticket_number or 'N/A'}")
        print(f"   描述: {email_info.get('short_description', 'N/A')}")
    
    def export_summary_to_excel(self, summary_data, title, filename):
        """将汇总数据导出为Excel文件"""
        import datetime
        
        if not summary_data:
            return False
        
        try:
            # 尝试导入openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                self.logger.error("需要安装openpyxl库: pip install openpyxl")
                return False
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "邮件处理汇总"
            
            # 设置标题
            ws['A1'] = title
            ws['A2'] = f"导出时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 标题样式
            title_font = Font(size=14, bold=True)
            ws['A1'].font = title_font
            ws['A2'].font = Font(size=10)
            
            # 检查是否有China Cloud邮件
            has_china_cloud = any(record.get('is_china_cloud', False) for record in summary_data)
            
            # 设置表头
            if has_china_cloud:
                headers = ['序号', '处理时间', '单号', 'Short Description', 'Requested by', 'Environment', 'Required Permissions', 'Reason']
                row_start = 4
            else:
                headers = ['序号', '处理时间', '单号', 'Short Description', 'Requested by']
                row_start = 4
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_start, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 写入数据
            for i, record in enumerate(summary_data, 1):
                row = row_start + i
                processed_time = record.get('processed_time', 'N/A')
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                # 写入基本列
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=processed_time)
                ws.cell(row=row, column=3, value=ticket_number)
                ws.cell(row=row, column=4, value=short_description)
                ws.cell(row=row, column=5, value=requested_by)
                
                # 如果有China Cloud字段，写入额外列
                if has_china_cloud:
                    if record.get('is_china_cloud', False):
                        environment = record.get('environment', 'N/A')
                        required = record.get('required_permissions', 'N/A') 
                        reason = record.get('reason_for_application', 'N/A')
                        
                        ws.cell(row=row, column=6, value=environment)
                        ws.cell(row=row, column=7, value=required)
                        ws.cell(row=row, column=8, value=reason)
                    else:
                        # 非China Cloud邮件
                        ws.cell(row=row, column=6, value="-")
                        ws.cell(row=row, column=7, value="-")
                        ws.cell(row=row, column=8, value="-")
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # 设置列宽，限制最大宽度为50
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 添加汇总行
            total_row = row_start + len(summary_data) + 2
            ws.cell(row=total_row, column=1, value=f"总计处理邮件: {len(summary_data)} 封")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            
            # 保存文件
            wb.save(filename)
            return True
            
        except Exception as e:
            self.logger.error(f"导出Excel文件失败: {e}")
            return False
    
    def export_summary_to_file(self, summary_data, title, filename, format_type='txt'):
        """将汇总数据输出到文件"""
        if format_type.lower() == 'xlsx':
            return self.export_summary_to_excel(summary_data, title, filename)
        else:
            return self.export_summary_to_txt(summary_data, title, filename)
    
    def export_summary_to_txt(self, summary_data, title, filename):
        """将汇总数据输出到文本文件"""
        import datetime
        
        if not summary_data:
            return False
        
        # 检查是否有China Cloud邮件
        has_china_cloud = any(record.get('is_china_cloud', False) for record in summary_data)
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                # 写入标题和时间戳
                f.write(f"{title}\n")
                f.write(f"导出时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                if has_china_cloud:
                    # 使用扩展格式 - 完整显示所有信息，不截断
                    f.write("=" * 350 + "\n")
                    f.write(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<60} {'Requested by':<30} {'Environment':<40} {'Required Permissions':<50} {'Reason':<40}\n")
                    f.write("-" * 350 + "\n")
                    
                    for i, record in enumerate(summary_data, 1):
                        processed_time = record.get('processed_time', 'N/A')[:19]
                        ticket_number = record.get('ticket_number', 'N/A')
                        short_description = record.get('short_description', 'N/A')
                        requested_by = record.get('requested_by', 'N/A')
                        
                        if record.get('is_china_cloud', False):
                            environment = record.get('environment', 'N/A')
                            required = record.get('required_permissions', 'N/A') 
                            reason = record.get('reason_for_application', 'N/A')
                            
                            f.write(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {environment:<40} {required:<50} {reason:<40}\n")
                        else:
                            # 非China Cloud邮件，其他列显示为"-"
                            f.write(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {'-':<40} {'-':<50} {'-':<40}\n")
                    
                    f.write("-" * 350 + "\n")
                else:
                    # 使用标准格式 - 完整显示，不截断
                    f.write("=" * 170 + "\n")
                    f.write(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<70} {'Requested by':<60}\n")
                    f.write("-" * 170 + "\n")
                    
                    for i, record in enumerate(summary_data, 1):
                        processed_time = record.get('processed_time', 'N/A')[:19]
                        ticket_number = record.get('ticket_number', 'N/A')
                        short_description = record.get('short_description', 'N/A')
                        requested_by = record.get('requested_by', 'N/A')
                        
                        f.write(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<70} {requested_by:<60}\n")
                    
                    f.write("-" * 170 + "\n")
                
                f.write(f"总计处理邮件: {len(summary_data)} 封\n")
                f.write("=" * (350 if has_china_cloud else 170) + "\n")
            
            return True
        except Exception as e:
            self.logger.error(f"导出文本文件失败: {e}")
            return False
    
    def print_processing_summary(self, export_to_file=False, export_format='txt'):
        """打印处理汇总报告"""
        summary_data = self.load_processing_summary()
        
        if not summary_data:
            print("📊 暂无处理记录")
            return
        
        # 导出到文件（如果需要）
        if export_to_file:
            import datetime
            if export_format.lower() == 'xlsx':
                filename = f"processing_summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            else:
                filename = f"processing_summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                
            if self.export_summary_to_file(summary_data, "📊 邮件处理汇总报告", filename, export_format):
                print(f"📁 汇总已导出到文件: {filename}")
            else:
                print("❌ 导出文件失败")
        
        # 检查是否有China Cloud邮件
        has_china_cloud = any(record.get('is_china_cloud', False) for record in summary_data)
        
        if has_china_cloud:
            # 使用扩展格式显示China Cloud邮件
            print("\n" + "=" * 300)
            print("📊 邮件处理汇总报告 (包含China Cloud详细信息)")
            print("=" * 300)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<60} {'Requested by':<30} {'Environment':<40} {'Required Permissions':<50} {'Reason':<40}")
            print("-" * 300)
            
            for i, record in enumerate(summary_data, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]  # 只显示到秒
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                if record.get('is_china_cloud', False):
                    environment = record.get('environment', 'N/A')
                    required = record.get('required_permissions', 'N/A') 
                    reason = record.get('reason_for_application', 'N/A')
                    
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {environment:<40} {required:<50} {reason:<40}")
                else:
                    # 非China Cloud邮件，其他列显示为"-"
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {'-':<40} {'-':<50} {'-':<40}")
            
            print("-" * 300)
            print(f"总计处理邮件: {len(summary_data)} 封")
            print("=" * 300)
        else:
            # 使用标准格式
            print("\n" + "=" * 170)
            print("📊 邮件处理汇总报告")
            print("=" * 170)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<70} {'Requested by':<60}")
            print("-" * 170)
            
            for i, record in enumerate(summary_data, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]  # 只显示到秒
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<70} {requested_by:<60}")
            
            print("-" * 170)
            print(f"总计处理邮件: {len(summary_data)} 封")
            print("=" * 170)
    
    def print_daily_processing_summary(self, target_date=None, export_to_file=False, export_format='txt'):
        """打印指定日期的处理汇总报告"""
        import datetime
        
        if target_date is None:
            target_date = datetime.date.today().strftime('%Y-%m-%d')
        
        summary_data = self.load_processing_summary()
        
        # 按日期过滤
        daily_records = []
        for record in summary_data:
            processed_time = record.get('processed_time', '')
            if processed_time.startswith(target_date):
                daily_records.append(record)
        
        if not daily_records:
            print(f"📊 {target_date} 暂无处理记录")
            return
        
        # 导出到文件（如果需要）
        if export_to_file:
            if export_format.lower() == 'xlsx':
                filename = f"daily_summary_{target_date}.xlsx"
            else:
                filename = f"daily_summary_{target_date}.txt"
            if self.export_summary_to_file(daily_records, f"📊 {target_date} 邮件处理汇总报告", filename, export_format):
                print(f"📁 {target_date} 汇总已导出到文件: {filename}")
            else:
                print("❌ 导出文件失败")
        
        # 检查是否有China Cloud邮件
        has_china_cloud = any(record.get('is_china_cloud', False) for record in daily_records)
        
        if has_china_cloud:
            # 使用扩展格式显示China Cloud邮件 - 增加列宽度
            print("\n" + "=" * 300)
            print(f"📊 {target_date} 邮件处理汇总报告 (包含China Cloud详细信息)")
            print("=" * 300)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<60} {'Requested by':<30} {'Environment':<40} {'Required Permissions':<50} {'Reason':<40}")
            print("-" * 300)
            
            for i, record in enumerate(daily_records, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]  # 只显示到秒
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                if record.get('is_china_cloud', False):
                    environment = record.get('environment', 'N/A')
                    required = record.get('required_permissions', 'N/A') 
                    reason = record.get('reason_for_application', 'N/A')
                    
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {environment:<40} {required:<50} {reason:<40}")
                else:
                    # 非China Cloud邮件，其他列显示为"-"
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {'-':<40} {'-':<50} {'-':<40}")
            
            print("-" * 300)
            print(f"{target_date} 处理邮件: {len(daily_records)} 封")
            print("=" * 300)
        else:
            # 使用标准格式
            print("\n" + "=" * 170)
            print(f"📊 {target_date} 邮件处理汇总报告")
            print("=" * 170)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<70} {'Requested by':<60}")
            print("-" * 170)
            
            for i, record in enumerate(daily_records, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]  # 只显示到秒
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<70} {requested_by:<60}")
            
            print("-" * 170)
            print(f"{target_date} 处理邮件: {len(daily_records)} 封")
            print("=" * 170)
    
    def parse_mbox_file(self, mbox_path):
        """解析Thunderbird的mbox文件"""
        emails = []
        
        try:
            with open(mbox_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # 按照mbox格式分割邮件 - 但跳过Mozilla头部
            lines = content.split('\n')
            current_email_lines = []
            emails_raw = []
            
            in_mozilla_headers = True
            found_real_headers = False
            
            for line in lines:
                # 检查是否是新邮件的开始（mbox分隔符）
                if line.startswith('From ') and not line.startswith('>>From '):
                    # 如果有当前邮件内容并且找到了真实邮件头，保存它
                    if current_email_lines and found_real_headers:
                        emails_raw.append('\n'.join(current_email_lines))
                    current_email_lines = []  # 不包含mbox分隔符行
                    in_mozilla_headers = True
                    found_real_headers = False
                    continue
                
                # 跳过Mozilla特定的头部
                if in_mozilla_headers and (
                    line.startswith('X-Mozilla-Status') or 
                    line.startswith('X-Mozilla-Keys') or
                    line.strip() == ''
                ):
                    continue
                
                # 一旦遇到标准邮件头，就不再是Mozilla头部了
                if in_mozilla_headers and (
                    line.startswith('Received:') or
                    line.startswith('From:') or
                    line.startswith('To:') or
                    line.startswith('Subject:') or
                    line.startswith('Date:')
                ):
                    in_mozilla_headers = False
                    found_real_headers = True
                
                # 添加到当前邮件内容（如果不在Mozilla头部）
                if not in_mozilla_headers:
                    current_email_lines.append(line)
            
            # 添加最后一封邮件
            if current_email_lines and found_real_headers:
                emails_raw.append('\n'.join(current_email_lines))
            
            # 解析每封邮件
            for raw_email in emails_raw:
                try:
                    msg = email.message_from_string(raw_email)
                    
                    email_info = {
                        'message_id': msg.get('Message-ID', ''),
                        'from': msg.get('From', ''),
                        'to': msg.get('To', ''),
                        'subject': msg.get('Subject', ''),
                        'date': msg.get('Date', ''),
                        'reply_to': msg.get('Reply-To') or msg.get('From', ''),
                        'file_path': mbox_path,
                        'raw_message': raw_email
                    }
                    
                    # 提取正文内容
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                try:
                                    body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                                    break
                                except:
                                    continue
                    else:
                        try:
                            payload = msg.get_payload()
                            if isinstance(payload, bytes):
                                body = payload.decode('utf-8', errors='ignore')
                            else:
                                body = str(payload)
                        except:
                            body = str(msg.get_payload())
                    
                    email_info['body'] = body
                    
                    # 提取Short description字段
                    short_description = self.extract_short_description(body)
                    email_info['short_description'] = short_description
                    
                    # 如果是China Cloud相关邮件，提取额外字段
                    if short_description and (short_description.lower().startswith('china cloud account and permission request') or 
                                             short_description.lower().startswith('china cloud resource request')):
                        if short_description.lower().startswith('china cloud account and permission request'):
                            china_cloud_fields = self.extract_china_cloud_fields(body)
                        else:  # China Cloud Resource Request
                            china_cloud_fields = self.extract_china_cloud_resource_fields(body)
                        email_info.update(china_cloud_fields)
                    
                    # 如果是CN-Server & DB Access Control邮件，提取额外字段
                    elif short_description and short_description.lower().startswith('cn-server & db access control'):
                        cn_server_fields = self.extract_cn_server_db_access_fields(body)
                        email_info.update(cn_server_fields)
                    
                    # 只添加有效的邮件（至少有主题或发件人）
                    if email_info['subject'] or email_info['from']:
                        emails.append(email_info)
                        
                except Exception as e:
                    self.logger.warning(f"解析邮件失败: {e}")
                    continue
                    
        except Exception as e:
            self.logger.error(f"解析mbox文件失败 {mbox_path}: {e}")
            
        return emails
    
    def extract_short_description(self, body):
        """从邮件正文中提取Short description字段"""
        import re
        
        if not body:
            return ""
        
        # 寻找Short description的模式，更精确的匹配
        patterns = [
            r'Short\s+description[:\s]*([^\r\n]+)',
            r'Short\s+Description[:\s]*([^\r\n]+)', 
            r'(?:^|\n)\s*Description[:\s]*([^\r\n]+)',
            r'(?:^|\n)\s*Request[:\s]*([^\r\n]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                description = match.group(1).strip()
                # 清理多余的空白字符和特殊字符
                description = re.sub(r'\s+', ' ', description)
                # 过滤掉太短或明显错误的匹配
                if len(description) > 3 and not description.startswith('字段'):
                    return description
        
        # 如果没找到，返回空字符串
        return ""
    
    def extract_china_cloud_fields(self, body):
        """从邮件正文中提取China Cloud相关字段"""
        import re
        
        if not body:
            return {}
        
        fields = {
            'environment': '',
            'required_permissions': '',
            'reason_for_application': '',
            'requested_by': ''
        }
        
        # 调试：记录正在处理的China Cloud邮件
        self.logger.info("正在提取China Cloud字段...")
        
        # 提取Requested by字段
        requested_by_patterns = [
            r'Requested\s+by[:\s]*([^\r\n]+)',
            r'Requested\s+for[:\s]*([^\r\n]+)',
            r'Request\s+by[:\s]*([^\r\n]+)',
            r'Request\s+for[:\s]*([^\r\n]+)'
        ]
        
        for pattern in requested_by_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['requested_by'] = match.group(1).strip()
                self.logger.info(f"提取到Requested by: {fields['requested_by']}")
                break
        
        if not fields['requested_by']:
            self.logger.warning("未找到Requested by字段")
        
        # 提取Environment (Permission regards to environment的值)
        env_patterns = [
            r'Permission\s+regards\s+to\s+environm?e?nt[:\s]*([^\r\n]+)',  # 支持 environmnt 或 environment
            r'Environment[:\s]*([^\r\n]+)',
            r'Regards\s+to\s+environm?e?nt[:\s]*([^\r\n]+)'  # 支持 environmnt 或 environment
        ]
        
        for pattern in env_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['environment'] = match.group(1).strip()
                self.logger.info(f"提取到Environment: {fields['environment']}")
                break
        
        if not fields['environment']:
            self.logger.warning("未找到Environment字段")
        
        # 提取Required permissions - 支持多行内容，使用与Resource Info相同的策略
        required_start_pattern = r'Required\s+permissions[:\s]*'
        required_match = re.search(required_start_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if required_match:
            start_pos = required_match.end()
            remaining_text = body[start_pos:]
            
            # 寻找下一个字段的开始
            next_field_patterns = [
                r'\n\s*Environment\s*:',
                r'\n\s*Reason\s+for\s+application\s*:',
                r'\n\s*Justification\s*:',
                r'\n\s*Notes\s*:',
                r'\n\s*Comments\s*:',
                r'\n\s*Additional\s+information\s*:',
                r'\n\s*Permission\s+regards\s+to\s+environment\s*:',
                r'\n\s*[A-Za-z]+\s+[A-Za-z]+\s*:'  # 通用的字段模式
            ]
            
            end_pos = len(remaining_text)
            for pattern in next_field_patterns:
                match = re.search(pattern, remaining_text, re.IGNORECASE | re.MULTILINE)
                if match:
                    end_pos = min(end_pos, match.start())
            
            required_text = remaining_text[:end_pos].strip()
            
            # 清理文本
            lines = required_text.split('\n')
            cleaned_lines = []
            for line in lines:
                line = line.strip()
                # 过滤掉空行、分隔符行和明显的格式化字符
                if (line and 
                    not line.startswith('=') and 
                    not line.startswith('-') and
                    not re.match(r'^[\s\-=_]+$', line) and
                    len(line.replace(' ', '').replace('\t', '')) > 0):
                    cleaned_lines.append(line)
            
            if cleaned_lines:
                required_text = ' '.join(cleaned_lines)
                required_text = re.sub(r'\s+', ' ', required_text)  # 压缩多个空格
                fields['required_permissions'] = required_text.strip()
                self.logger.info(f"提取到Required Permissions (多行策略): {fields['required_permissions']}")
        else:
            # 备用简单提取方法
            perm_patterns = [
                r'Required\s+permissions[:\s]*([^\r\n]+)',
                r'Permissions\s+required[:\s]*([^\r\n]+)',
                r'Permission[:\s]*([^\r\n]+)',
                r'Required\s+permission[:\s]*([^\r\n]+)'  # 单数形式
            ]
            
            for pattern in perm_patterns:
                match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
                if match:
                    fields['required_permissions'] = match.group(1).strip()
                    self.logger.info(f"提取到Required Permissions (单行备用): {fields['required_permissions']}")
                    break
        
        if not fields['required_permissions']:
            self.logger.warning("未找到Required Permissions字段")
        
        # 提取Reason for application
        reason_patterns = [
            r'Reason\s+for\s+application[:\s]*([^\r\n]+)',
            r'Application\s+reason[:\s]*([^\r\n]+)',
            r'Reason[:\s]*([^\r\n]+)',
            r'Justification[:\s]*([^\r\n]+)'  # 添加另一种可能的字段名
        ]
        
        for pattern in reason_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['reason_for_application'] = match.group(1).strip()
                self.logger.info(f"提取到Reason: {fields['reason_for_application']}")
                break
        
        if not fields['reason_for_application']:
            self.logger.warning("未找到Reason for Application字段")
        
        # 调试：如果所有字段都为空，输出邮件正文片段供调试
        if not any(fields.values()):
            self.logger.warning("所有China Cloud字段都为空，邮件正文片段:")
            # 输出前500个字符用于调试
            debug_text = body[:500].replace('\n', '\\n').replace('\r', '\\r')
            self.logger.warning(f"邮件正文: {debug_text}")
        
        return fields
    
    def extract_china_cloud_resource_fields(self, body):
        """从邮件正文中提取China Cloud Resource Request相关字段"""
        import re
        
        if not body:
            return {}
        
        fields = {
            'environment': '',
            'required_permissions': '',
            'reason_for_application': '',
            'requested_by': ''
        }
        
        # 调试：记录正在处理的China Cloud Resource Request邮件
        self.logger.info("正在提取China Cloud Resource Request字段...")
        
        # 提取Requested by字段
        requested_by_patterns = [
            r'Requested\s+by[:\s]*([^\r\n]+)',
            r'Requested\s+for[:\s]*([^\r\n]+)',
            r'Request\s+by[:\s]*([^\r\n]+)',
            r'Request\s+for[:\s]*([^\r\n]+)'
        ]
        
        for pattern in requested_by_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['requested_by'] = match.group(1).strip()
                self.logger.info(f"提取到Requested by: {fields['requested_by']}")
                break
        
        if not fields['requested_by']:
            self.logger.warning("未找到Requested by字段")
        
        # 提取Environment字段
        env_patterns = [
            r'Environment[:\s]*([^\r\n]+)',
            r'Env[:\s]*([^\r\n]+)',
            r'Environment\s*[:\s]*([^\r\n]+)'
        ]
        
        for pattern in env_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['environment'] = match.group(1).strip()
                self.logger.info(f"提取到Environment: {fields['environment']}")
                break
        
        if not fields['environment']:
            self.logger.warning("未找到Environment字段")
        
        # 提取Resource Info作为Required permissions - 支持多行内容，更宽泛的策略
        resource_start_pattern = r'Resource\s+Info[:\s]*'
        resource_match = re.search(resource_start_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if resource_match:
            start_pos = resource_match.end()
            remaining_text = body[start_pos:]
            
            # 寻找下一个字段的开始（Environment, Reason for application等）
            next_field_patterns = [
                r'\n\s*Environment\s*:',
                r'\n\s*Reason\s+for\s+application\s*:',
                r'\n\s*Justification\s*:',
                r'\n\s*Notes\s*:',
                r'\n\s*Comments\s*:',
                r'\n\s*Additional\s+information\s*:',
                r'\n\s*[A-Za-z]+\s+[A-Za-z]+\s*:'  # 通用的字段模式
            ]
            
            end_pos = len(remaining_text)
            for pattern in next_field_patterns:
                match = re.search(pattern, remaining_text, re.IGNORECASE | re.MULTILINE)
                if match:
                    end_pos = min(end_pos, match.start())
            
            resource_text = remaining_text[:end_pos].strip()
            
            # 清理文本
            lines = resource_text.split('\n')
            cleaned_lines = []
            for line in lines:
                line = line.strip()
                # 过滤掉空行、分隔符行和明显的格式化字符
                if (line and 
                    not line.startswith('=') and 
                    not line.startswith('-') and
                    not re.match(r'^[\s\-=_]+$', line) and
                    len(line.replace(' ', '').replace('\t', '')) > 0):
                    cleaned_lines.append(line)
            
            if cleaned_lines:
                resource_text = ' '.join(cleaned_lines)
                resource_text = re.sub(r'\s+', ' ', resource_text)  # 压缩多个空格
                fields['required_permissions'] = resource_text.strip()
                self.logger.info(f"提取到Resource Info (多行策略): {fields['required_permissions']}")
        else:
            # 备用简单提取方法
            resource_patterns = [
                r'Resource\s+Info[:\s]*([^\r\n]+)',
                r'Resource\s+Information[:\s]*([^\r\n]+)',
                r'Resource[:\s]*([^\r\n]+)',
                r'Resources[:\s]*([^\r\n]+)'
            ]
            
            for pattern in resource_patterns:
                match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
                if match:
                    fields['required_permissions'] = match.group(1).strip()
                    self.logger.info(f"提取到Resource Info (单行备用): {fields['required_permissions']}")
                    break
        
        if not fields['required_permissions']:
            self.logger.warning("未找到Resource Info字段")
        
        # 提取Reason for application
        reason_patterns = [
            r'Reason\s+for\s+application[:\s]*([^\r\n]+)',
            r'Application\s+reason[:\s]*([^\r\n]+)',
            r'Reason[:\s]*([^\r\n]+)',
            r'Justification[:\s]*([^\r\n]+)'
        ]
        
        for pattern in reason_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['reason_for_application'] = match.group(1).strip()
                self.logger.info(f"提取到Reason: {fields['reason_for_application']}")
                break
        
        if not fields['reason_for_application']:
            self.logger.warning("未找到Reason for Application字段")
        
        # 调试：如果所有字段都为空，输出邮件正文片段供调试
        if not any(fields.values()):
            self.logger.warning("所有China Cloud Resource Request字段都为空，邮件正文片段:")
            # 输出前500个字符用于调试
            debug_text = body[:500].replace('\n', '\\n').replace('\r', '\\r')
            self.logger.warning(f"邮件正文: {debug_text}")
        
        return fields
    
    def extract_cn_server_db_access_fields(self, body):
        """从邮件正文中提取CN-Server & DB Access Control相关字段"""
        import re
        
        if not body:
            return {}
        
        fields = {
            'environment': '',
            'required_permissions': '',
            'reason_for_application': '',
            'requested_by': ''
        }
        
        # 调试：记录正在处理的CN-Server & DB Access Control邮件
        self.logger.info("正在提取CN-Server & DB Access Control字段...")
        
        # 提取Requested by字段
        requested_by_patterns = [
            r'Requested\s+by[:\s]*([^\r\n]+)',
            r'Requested\s+for[:\s]*([^\r\n]+)',
            r'Request\s+by[:\s]*([^\r\n]+)',
            r'Request\s+for[:\s]*([^\r\n]+)'
        ]
        
        for pattern in requested_by_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['requested_by'] = match.group(1).strip()
                self.logger.info(f"提取到Requested by: {fields['requested_by']}")
                break
        
        if not fields['requested_by']:
            self.logger.warning("未找到Requested by字段")
        
        # 首先检查"What System do you need access to?"是否为Bastion
        system_access_pattern = r'What\s+System\s+do\s+you\s+need\s+access\s+to\?[:\s]*([^\r\n]+)'
        system_match = re.search(system_access_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if system_match:
            system_value = system_match.group(1).strip()
            self.logger.info(f"找到System access值: {system_value}")
            
            if system_value.lower() != 'bastion':
                self.logger.info(f"System access不是Bastion，跳过CN-Server & DB Access Control字段提取")
                return fields
        else:
            self.logger.warning("未找到'What System do you need access to?'字段")
            return fields
        
        # 提取Environment字段
        env_patterns = [
            r'Environment[:\s]*([^\r\n]+)',
            r'Env[:\s]*([^\r\n]+)',
            r'Environment\s*[:\s]*([^\r\n]+)'
        ]
        
        for pattern in env_patterns:
            match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
            if match:
                fields['environment'] = match.group(1).strip()
                self.logger.info(f"提取到Environment: {fields['environment']}")
                break
        
        if not fields['environment']:
            self.logger.warning("未找到Environment字段")
        
        # 提取Authorization time作为Required permissions - 支持多行内容
        auth_time_start_pattern = r'Authorization\s+time[:\s]*'
        auth_time_match = re.search(auth_time_start_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if auth_time_match:
            start_pos = auth_time_match.end()
            remaining_text = body[start_pos:]
            
            # 寻找下一个字段的开始
            next_field_patterns = [
                r'\n\s*Environment\s*:',
                r'\n\s*Reason\s+for\s+application\s*:',
                r'\n\s*What\s+System\s*:',
                r'\n\s*Justification\s*:',
                r'\n\s*Notes\s*:',
                r'\n\s*Comments\s*:',
                r'\n\s*[A-Za-z]+\s+[A-Za-z]+\s*:'  # 通用的字段模式
            ]
            
            end_pos = len(remaining_text)
            for pattern in next_field_patterns:
                match = re.search(pattern, remaining_text, re.IGNORECASE | re.MULTILINE)
                if match:
                    end_pos = min(end_pos, match.start())
            
            auth_time_text = remaining_text[:end_pos].strip()
            
            # 清理文本
            lines = auth_time_text.split('\n')
            cleaned_lines = []
            for line in lines:
                line = line.strip()
                # 过滤掉空行、分隔符行和明显的格式化字符
                if (line and 
                    not line.startswith('=') and 
                    not line.startswith('-') and
                    not re.match(r'^[\s\-=_]+$', line) and
                    len(line.replace(' ', '').replace('\t', '')) > 0):
                    cleaned_lines.append(line)
            
            if cleaned_lines:
                auth_time_text = ' '.join(cleaned_lines)
                auth_time_text = re.sub(r'\s+', ' ', auth_time_text)  # 压缩多个空格
                fields['required_permissions'] = auth_time_text.strip()
                self.logger.info(f"提取到Authorization time (多行策略): {fields['required_permissions']}")
        else:
            # 备用简单提取方法
            auth_time_patterns = [
                r'Authorization\s+time[:\s]*([^\r\n]+)',
                r'Auth\s+time[:\s]*([^\r\n]+)',
                r'Authorization[:\s]*([^\r\n]+)'
            ]
            
            for pattern in auth_time_patterns:
                match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
                if match:
                    fields['required_permissions'] = match.group(1).strip()
                    self.logger.info(f"提取到Authorization time (单行备用): {fields['required_permissions']}")
                    break
        
        if not fields['required_permissions']:
            self.logger.warning("未找到Authorization time字段")
        
        # 提取Reason for application (including reason for Authorization time)
        reason_start_pattern = r'Reason\s+for\s+application\s*\(including\s+reason\s+for\s+Authorization\s+time\)[:\s]*'
        reason_match = re.search(reason_start_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if not reason_match:
            # 尝试更简单的模式
            reason_start_pattern = r'Reason\s+for\s+application[:\s]*'
            reason_match = re.search(reason_start_pattern, body, re.IGNORECASE | re.MULTILINE)
        
        if reason_match:
            start_pos = reason_match.end()
            remaining_text = body[start_pos:]
            
            # 寻找下一个字段的开始或文本结尾
            next_field_patterns = [
                r'\n\s*Environment\s*:',
                r'\n\s*Authorization\s+time\s*:',
                r'\n\s*What\s+System\s*:',
                r'\n\s*Justification\s*:',
                r'\n\s*Notes\s*:',
                r'\n\s*Comments\s*:',
                r'\n\s*[A-Za-z]+\s+[A-Za-z]+\s*:'  # 通用的字段模式
            ]
            
            end_pos = len(remaining_text)
            for pattern in next_field_patterns:
                match = re.search(pattern, remaining_text, re.IGNORECASE | re.MULTILINE)
                if match:
                    end_pos = min(end_pos, match.start())
            
            reason_text = remaining_text[:end_pos].strip()
            
            # 清理文本
            lines = reason_text.split('\n')
            cleaned_lines = []
            for line in lines:
                line = line.strip()
                # 过滤掉空行、分隔符行和明显的格式化字符
                if (line and 
                    not line.startswith('=') and 
                    not line.startswith('-') and
                    not re.match(r'^[\s\-=_]+$', line) and
                    len(line.replace(' ', '').replace('\t', '')) > 0):
                    cleaned_lines.append(line)
            
            if cleaned_lines:
                reason_text = ' '.join(cleaned_lines)
                reason_text = re.sub(r'\s+', ' ', reason_text)  # 压缩多个空格
                fields['reason_for_application'] = reason_text.strip()
                self.logger.info(f"提取到Reason (多行策略): {fields['reason_for_application']}")
        else:
            # 备用简单提取方法
            reason_patterns = [
                r'Reason\s+for\s+application[:\s]*([^\r\n]+)',
                r'Application\s+reason[:\s]*([^\r\n]+)',
                r'Reason[:\s]*([^\r\n]+)',
                r'Justification[:\s]*([^\r\n]+)'
            ]
            
            for pattern in reason_patterns:
                match = re.search(pattern, body, re.IGNORECASE | re.MULTILINE)
                if match:
                    fields['reason_for_application'] = match.group(1).strip()
                    self.logger.info(f"提取到Reason (单行备用): {fields['reason_for_application']}")
                    break
        
        if not fields['reason_for_application']:
            self.logger.warning("未找到Reason for application字段")
        
        # 调试：如果所有字段都为空，输出邮件正文片段供调试
        if not any(fields.values()):
            self.logger.warning("所有CN-Server & DB Access Control字段都为空，邮件正文片段:")
            # 输出前500个字符用于调试
            debug_text = body[:500].replace('\n', '\\n').replace('\r', '\\r')
            self.logger.warning(f"邮件正文: {debug_text}")
        
        return fields
    
    def is_approval_needed(self, email_info):
        """判断是否需要批准"""
        subject = email_info.get('subject', '').lower()
        
        # 检查是否包含需要批准的关键词
        approval_keywords = ['approve', 'approval', 'ritm', 'servicenow']
        
        for keyword in approval_keywords:
            if keyword in subject:
                return True
                
        return False
    
    def create_approval_reply(self, original_email):
        """创建批准回复邮件"""
        try:
            msg = MIMEMultipart('alternative')
            
            # 设置邮件头
            from_name = self.config.get('EMAIL', 'from_name')
            from_email = self.config.get('EMAIL', 'from_email')
            msg['From'] = f"{from_name} <{from_email}>"
            msg['To'] = original_email['reply_to']
            
            # 生成回复主题 - 提取RITM号码并格式化为标准格式
            original_subject = original_email.get('subject', '')
            
            # 从原主题中提取RITM号码
            import re
            ritm_match = re.search(r'RITM(\d+)', original_subject, re.IGNORECASE)
            
            if ritm_match:
                ritm_number = ritm_match.group(1)
                # 确保格式化为标准回复主题，始终包含Re:前缀
                reply_subject = f"Re: RITM{ritm_number} - approve"
                msg['Subject'] = reply_subject
                self.logger.info(f"生成回复主题: {reply_subject}")
            else:
                # 如果找不到RITM号码，使用原始逻辑，但确保有Re:前缀
                if not original_subject.lower().startswith('re:'):
                    reply_subject = f"Re: {original_subject}"
                else:
                    reply_subject = original_subject
                msg['Subject'] = reply_subject
                self.logger.warning(f"未找到RITM号码，使用备用主题: {reply_subject}")
                
            msg['Date'] = formatdate(localtime=True)
            msg['Message-ID'] = make_msgid()
            
            # 从原始邮件内容中提取Ref信息
            original_body = original_email.get('body', '')
            
            # 搜索Ref信息，获取最后出现的MSG号码
            import re
            msg_matches = re.findall(r'MSG\d+', original_body)
            
            if msg_matches:
                # 使用最后出现的MSG号码
                ref_msg = f"Ref:{msg_matches[-1]}"
                approval_msg = ref_msg
                self.logger.info(f"从原始邮件提取到Ref信息: {approval_msg}")
            else:
                # 备用方案：使用配置文件中的默认值
                approval_msg = self.config.get('EMAIL', 'approval_message')
                self.logger.warning(f"未能从邮件中提取Ref信息，使用默认值: {approval_msg}")
            
            # 纯文本部分 - 简洁版本
            text_content = approval_msg
            
            # HTML部分 - 简洁版本
            html_content = f"""<html>
<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">
</head>
<body>
{approval_msg}
</body>
</html>"""
            
            # 添加纯文本和HTML部分
            text_part = MIMEText(text_content, 'plain', 'utf-8')
            html_part = MIMEText(html_content, 'html', 'utf-8')
            
            msg.attach(text_part)
            msg.attach(html_part)
            
            return msg
            
        except Exception as e:
            self.logger.error(f"创建回复邮件失败: {e}")
            return None
    
    def move_processed_email(self, email_info, source_mbox_path):
        """将已处理的邮件移动到目标文件夹"""
        try:
            # 检查是否启用邮件移动
            if not self.config.getboolean('DEFAULT', 'move_processed_emails', fallback=False):
                return True
            
            processed_destination = self.config.get('DEFAULT', 'processed_destination', fallback='')
            if not processed_destination:
                self.logger.warning("未配置processed_destination，跳过邮件移动")
                return True
            
            # 查找目标文件夹路径
            target_path = self.find_thunderbird_mail_folder(processed_destination)
            if not target_path:
                self.logger.warning(f"找不到目标文件夹: {processed_destination}，尝试创建...")
                
                # 尝试创建目标文件夹路径
                if self._create_target_folder(processed_destination):
                    target_path = self.find_thunderbird_mail_folder(processed_destination)
                    if not target_path:
                        self.logger.error("创建目标文件夹后仍然找不到路径")
                        return False
                else:
                    return False
            
            # 获取目标mbox文件路径
            folder_name = processed_destination.split('/')[-1]
            target_mbox = os.path.join(target_path, folder_name)
            
            # 移动邮件：将邮件内容追加到目标文件，然后从源文件删除
            return self._move_email_between_mbox(email_info, source_mbox_path, target_mbox)
            
        except Exception as e:
            self.logger.error(f"移动邮件失败: {e}")
            return False
    
    def _create_target_folder(self, folder_path):
        """创建目标文件夹结构"""
        try:
            # 解析文件夹路径
            path_parts = folder_path.split('/')
            profile_path = self.config.get('DEFAULT', 'thunderbird_profile_path')
            
            # 查找profile目录
            for item in os.listdir(profile_path):
                profile_dir = os.path.join(profile_path, item)
                if os.path.isdir(profile_dir):
                    mail_dir = os.path.join(profile_dir, 'Mail')
                    if os.path.exists(mail_dir):
                        current_path = mail_dir
                        break
            else:
                return False
            
            # 构建文件夹结构
            if path_parts[0] == 'webaccountMail':
                # IMAP路径
                current_path = os.path.join(current_path, path_parts[0], path_parts[1])
                if not os.path.exists(current_path):
                    self.logger.error(f"IMAP账户路径不存在: {current_path}")
                    return False
                path_parts = path_parts[2:]  # 跳过webaccountMail/server部分
            elif path_parts[0] == 'Local':
                # Local Folders路径
                current_path = os.path.join(current_path, 'Local Folders')
                path_parts = path_parts[2:]  # 跳过'Local Folders'
            
            # 创建剩余的文件夹结构
            for i, folder in enumerate(path_parts):
                if i == len(path_parts) - 1:
                    # 最后一级，创建mbox文件
                    mbox_file = os.path.join(current_path, folder)
                    if not os.path.exists(mbox_file):
                        # 创建空的mbox文件
                        with open(mbox_file, 'w', encoding='utf-8') as f:
                            f.write('')
                        self.logger.info(f"创建了mbox文件: {mbox_file}")
                else:
                    # 中间级别，创建.sbd目录
                    folder_file = os.path.join(current_path, folder)
                    folder_sbd = folder_file + '.sbd'
                    
                    if not os.path.exists(folder_file):
                        with open(folder_file, 'w', encoding='utf-8') as f:
                            f.write('')
                    
                    if not os.path.exists(folder_sbd):
                        os.makedirs(folder_sbd)
                        self.logger.info(f"创建了文件夹: {folder_sbd}")
                    
                    current_path = folder_sbd
            
            return True
            
        except Exception as e:
            self.logger.error(f"创建目标文件夹失败: {e}")
            return False
    
    def _move_email_between_mbox(self, email_info, source_mbox, target_mbox):
        """在mbox文件之间移动单封邮件"""
        try:
            raw_message = email_info.get('raw_message', '')
            if not raw_message:
                self.logger.error("邮件内容为空，无法移动")
                return False
            
            # 确保目标文件存在
            if not os.path.exists(target_mbox):
                with open(target_mbox, 'w', encoding='utf-8') as f:
                    f.write('')
            
            # 添加到目标文件
            with open(target_mbox, 'a', encoding='utf-8', errors='ignore') as f:
                # 添加mbox分隔符
                if os.path.getsize(target_mbox) > 0:
                    f.write('\n')
                import time
                f.write(f"From - {time.strftime('%a %b %d %Y %H:%M:%S')}\n")
                f.write(raw_message)
                if not raw_message.endswith('\n'):
                    f.write('\n')
            
            # 从源文件删除（通过重写文件，排除当前邮件）
            self._remove_email_from_mbox(email_info, source_mbox)
            
            self.logger.info(f"✅ 邮件已移动到Processed: {email_info.get('subject', 'N/A')}")
            return True
            
        except Exception as e:
            self.logger.error(f"移动邮件内容失败: {e}")
            return False
    
    def _remove_email_from_mbox(self, email_to_remove, mbox_path):
        """从mbox文件中删除指定邮件"""
        try:
            # 重新解析mbox文件，排除要删除的邮件
            all_emails = self.parse_mbox_file(mbox_path)
            
            # 过滤掉要删除的邮件
            remaining_emails = []
            target_message_id = email_to_remove.get('message_id', '')
            target_subject = email_to_remove.get('subject', '')
            target_from = email_to_remove.get('from', '')
            
            for email in all_emails:
                # 使用多个字段进行匹配以确保准确性
                if (email.get('message_id', '') == target_message_id and target_message_id) or \
                   (email.get('subject', '') == target_subject and email.get('from', '') == target_from):
                    continue  # 跳过要删除的邮件
                remaining_emails.append(email)
            
            # 重写mbox文件
            with open(mbox_path, 'w', encoding='utf-8', errors='ignore') as f:
                for i, email in enumerate(remaining_emails):
                    if i > 0:
                        f.write('\n')
                    import time
                    f.write(f"From - {time.strftime('%a %b %d %Y %H:%M:%S')}\n")
                    f.write(email.get('raw_message', ''))
                    if not email.get('raw_message', '').endswith('\n'):
                        f.write('\n')
            
            self.logger.info(f"从源文件删除邮件: {target_subject}")
            return True
            
        except Exception as e:
            self.logger.error(f"从mbox删除邮件失败: {e}")
            return False
    
    def send_reply_via_email_client(self, reply_msg, original_email):
        """通过配置的邮件客户端发送回复邮件"""
        try:
            # 获取邮件客户端配置
            email_client = self.config.get('DEFAULT', 'email_client', fallback='thunderbird')
            
            # 提取邮件信息
            to_addr = reply_msg['To']
            subject = reply_msg['Subject']
            
            # 提取纯文本内容
            body_text = ""
            if reply_msg.is_multipart():
                for part in reply_msg.walk():
                    if part.get_content_type() == "text/plain":
                        try:
                            payload = part.get_payload(decode=True)
                            if payload:
                                body_text = payload.decode('utf-8', errors='ignore')
                                break
                        except:
                            continue
            
            if not body_text:
                # 备用：从配置生成简单正文
                approval_msg = self.config.get('EMAIL', 'approval_message')
                body_text = approval_msg
            
            # 检查发送模式
            send_mode = self.config.get('DEFAULT', 'send_mode', fallback='auto')
            auto_send = (send_mode.lower() == 'auto')
            
            self.logger.info(f"准备发送邮件到: {to_addr}")
            self.logger.info(f"主题: {subject}")
            self.logger.info(f"邮件客户端: {email_client}")
            self.logger.info(f"发送模式: {'自动发送' if auto_send else '只准备，需手动发送'}")
            
            # 根据配置选择邮件客户端
            if email_client.lower() == 'outlook':
                # 使用Outlook PWA发送
                from outlook_sender import OutlookPWASender
                browser = self.config.get('DEFAULT', 'outlook_browser', fallback='Safari')
                sender = OutlookPWASender(browser=browser)
                success = sender.send_email(to_addr, subject, body_text, auto_send=auto_send)
            else:
                # 使用Thunderbird发送 (默认)
                from thunderbird_sender import ThunderbirdSender
                sender = ThunderbirdSender()
                success = sender.send_email_via_applescript_with_confirmation(
                    to_addr, subject, body_text, auto_send=auto_send
                )
            
            if success:
                if auto_send:
                    self.logger.info("✅ 邮件已自动发送")
                else:
                    self.logger.info("✅ 邮件已准备就绪，请手动点击发送按钮")
                return True
            else:
                self.logger.warning("⚠️ 邮件发送失败")
                return False
                
        except Exception as e:
            self.logger.error(f"发送回复邮件失败: {e}")
            
            # 备用方案：保存为草稿文件
            try:
                import time
                draft_file = f"approval_reply_{int(time.time())}.eml"
                with open(draft_file, 'w', encoding='utf-8') as f:
                    f.write(reply_msg.as_string())
                self.logger.info(f"已保存邮件草稿: {draft_file}")
                return True
            except Exception as e2:
                self.logger.error(f"保存草稿文件也失败: {e2}")
                return False
    
    def process_mbox_file(self, mbox_path, show_daily_summary=False):
        """处理mbox邮件文件"""
        try:
            self.logger.info(f"处理mbox文件: {mbox_path}")
            
            # 解析mbox文件中的所有邮件
            emails = self.parse_mbox_file(mbox_path)
            
            if not emails:
                self.logger.info("mbox文件中没有找到邮件")
                return {"processed_count": 0, "processed_emails": []}
            
            processed_emails = self.load_processed_emails()
            processed_count = 0
            current_batch_processed = []  # 记录本次处理的邮件
            
            for email_info in emails:
                # 检查是否需要批准
                if not self.is_approval_needed(email_info):
                    continue
                
                # 检查是否已经处理过
                message_id = email_info.get('message_id', '')
                if not message_id:
                    # 如果没有Message-ID，使用主题和发件人作为标识
                    message_id = f"{email_info.get('subject', '')}|{email_info.get('from', '')}"
                
                if message_id in processed_emails:
                    continue
                
                # 检查是否启用自动批准
                if not self.config.getboolean('DEFAULT', 'auto_approve_enabled', fallback=True):
                    self.logger.info("自动批准功能已禁用")
                    continue
                
                # 如果不是第一封邮件，等待足够时间确保前一封邮件处理完成
                if processed_count > 0:
                    wait_time = 15  # 等待15秒确保前一封邮件完全处理完成
                    self.logger.info(f"等待 {wait_time} 秒，确保前一封邮件处理完成...")
                    import time
                    time.sleep(wait_time)
                
                self.logger.info(f"处理需要批准的邮件: {email_info['subject']}")
                self.logger.info(f"发件人: {email_info['from']}")
                
                # 记录Short description
                short_description = email_info.get('short_description', '')
                if short_description:
                    self.logger.info(f"Short description: {short_description}")
                    print(f"📋 Short description: {short_description}")
                else:
                    self.logger.warning("未找到Short description字段")
                    print("⚠️ 未找到Short description字段")
                
                # 创建回复邮件
                reply_msg = self.create_approval_reply(email_info)
                if not reply_msg:
                    continue
                
                # 发送回复邮件
                if self.send_reply_via_email_client(reply_msg, email_info):
                    # 记录已处理的邮件
                    processed_emails.add(message_id)
                    processed_count += 1
                    self.logger.info(f"成功处理邮件 {processed_count}: {email_info['subject']}")
                    
                    # 添加到处理汇总
                    import time
                    processed_time = time.strftime('%Y-%m-%d %H:%M:%S')
                    import re
                    subject = email_info.get('subject', '')
                    ticket_match = re.search(r'(RITM\d+|CHG\d+)', subject, re.IGNORECASE)
                    ticket_number = ticket_match.group(1) if ticket_match else None
                    self.add_to_processing_summary(email_info, processed_time, ticket_number)
                    
                    # 记录本次处理的邮件信息
                    batch_record = {
                        'ticket_number': ticket_number or 'N/A',
                        'short_description': email_info.get('short_description', 'N/A'),
                        'processed_time': processed_time,
                        'subject': email_info.get('subject', 'N/A'),
                        'requested_by': email_info.get('requested_by', 'N/A')
                    }
                    
                    # 如果是China Cloud或CN-Server & DB Access Control邮件，添加额外字段
                    short_desc_lower = email_info.get('short_description', '').lower()
                    if (short_desc_lower.startswith('china cloud account and permission request') or 
                        short_desc_lower.startswith('china cloud resource request') or
                        short_desc_lower.startswith('cn-server & db access control')):
                        batch_record.update({
                            'environment': email_info.get('environment', 'N/A'),
                            'required_permissions': email_info.get('required_permissions', 'N/A'),
                            'reason_for_application': email_info.get('reason_for_application', 'N/A'),
                            'is_china_cloud': True
                        })
                    else:
                        batch_record['is_china_cloud'] = False
                    
                    current_batch_processed.append(batch_record)
                    
                    # 立即保存已处理的邮件记录，防止意外中断时丢失进度
                    self.save_processed_emails(processed_emails)
                    
                    # 移动已处理的邮件到Processed文件夹
                    move_success = self.move_processed_email(email_info, mbox_path)
                    if move_success:
                        self.logger.info(f"📁 邮件已移动到Processed文件夹")
                    else:
                        self.logger.warning(f"⚠️ 邮件处理成功，但移动到Processed文件夹失败")
                else:
                    self.logger.error(f"发送回复失败: {email_info['subject']}")
                    # 发送失败时也等待一下，避免快速重试造成更多问题
                    import time
                    time.sleep(5)
            
            if processed_count > 0:
                self.save_processed_emails(processed_emails)
                self.logger.info(f"本次处理了 {processed_count} 封邮件")
            
            # 返回处理结果和本次处理的邮件列表
            return {
                "processed_count": processed_count,
                "processed_emails": current_batch_processed
            }
            
        except Exception as e:
            self.logger.error(f"处理mbox文件失败 {mbox_path}: {e}")
            return {"processed_count": 0, "processed_emails": []}
    
    def print_current_batch_summary(self, processed_emails_list, export_to_file=False, export_format='txt'):
        """显示本次处理的邮件汇总"""
        if not processed_emails_list:
            print("📊 本次没有处理任何邮件")
            return
        
        # 导出到文件（如果需要）
        if export_to_file:
            import datetime
            if export_format.lower() == 'xlsx':
                filename = f"batch_summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            else:
                filename = f"batch_summary_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            if self.export_summary_to_file(processed_emails_list, "📊 本次处理邮件汇总", filename, export_format):
                print(f"📁 本次处理汇总已导出到文件: {filename}")
            else:
                print("❌ 导出文件失败")
        
        # 检查是否有China Cloud邮件
        has_china_cloud = any(record.get('is_china_cloud', False) for record in processed_emails_list)
        
        if has_china_cloud:
            # 使用扩展格式 - 进一步增加列宽度，减少截断
            print("\n" + "=" * 300)
            print("📊 本次处理邮件汇总 (包含China Cloud详细信息)")
            print("=" * 300)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<60} {'Requested by':<30} {'Environment':<40} {'Required Permissions':<50} {'Reason':<40}")
            print("-" * 300)
            
            for i, record in enumerate(processed_emails_list, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                if record.get('is_china_cloud', False):
                    environment = record.get('environment', 'N/A')
                    required = record.get('required_permissions', 'N/A') 
                    reason = record.get('reason_for_application', 'N/A')
                    
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {environment:<40} {required:<50} {reason:<40}")
                else:
                    # 非China Cloud邮件，其他列显示为"-"
                    print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<60} {requested_by:<30} {'-':<40} {'-':<50} {'-':<40}")
            
            print("-" * 300)
        else:
            # 使用标准格式 - 显示完整信息，不截断
            print("\n" + "=" * 170)
            print("📊 本次处理邮件汇总")
            print("=" * 170)
            print(f"{'序号':<4} {'处理时间':<20} {'单号':<15} {'Short Description':<70} {'Requested by':<60}")
            print("-" * 170)
            
            for i, record in enumerate(processed_emails_list, 1):
                processed_time = record.get('processed_time', 'N/A')[:19]
                ticket_number = record.get('ticket_number', 'N/A')
                short_description = record.get('short_description', 'N/A')
                requested_by = record.get('requested_by', 'N/A')
                
                print(f"{i:<4} {processed_time:<20} {ticket_number:<15} {short_description:<70} {requested_by:<60}")
            
            print("-" * 170)
        
        print(f"本次处理邮件总数: {len(processed_emails_list)} 封")
        print("=" * (300 if has_china_cloud else 170))
    
    def run_once_mode(self, export_to_file=False, export_format='txt'):
        """一次性处理模式 - 处理现有邮件并显示本次处理汇总"""
        print("🚀 启动一次性邮件处理模式...")
        
        # 获取监控路径
        watch_folder = self.config.get('DEFAULT', 'watch_folder')
        
        # 查找Thunderbird邮件文件夹
        full_watch_path = self.find_thunderbird_mail_folder(watch_folder)
        
        if not full_watch_path:
            print(f"错误: 找不到Thunderbird邮件文件夹: {watch_folder}")
            return False
        
        print(f"📁 找到邮件目录: {full_watch_path}")
        
        # 获取目标文件夹名
        folder_name = watch_folder.split('/')[-1]  # 获取最后一级文件夹名
        
        # 处理现有邮件
        print("📧 开始处理邮件...")
        mbox_file = os.path.join(full_watch_path, folder_name)
        
        if os.path.exists(mbox_file) and os.path.isfile(mbox_file):
            # 不显示全部汇总，而是准备显示本次处理的汇总
            result = self.process_mbox_file(mbox_file, show_daily_summary=False)
            
            # 检查返回的结果
            processed_count = result.get("processed_count", 0)
            processed_emails_list = result.get("processed_emails", [])
            
            if processed_count > 0:
                print("🎉 邮件处理完成！")
                # 显示本次处理的汇总
                self.print_current_batch_summary(processed_emails_list, export_to_file=export_to_file, export_format=export_format)
                return True
            else:
                print("ℹ️ 没有需要处理的新邮件")
                return False
        else:
            print(f"❌ 邮件文件不存在: {mbox_file}")
            return False

class EmailWatcher(FileSystemEventHandler):
    """邮件文件监控器"""
    
    def __init__(self, approver, watch_folder):
        self.approver = approver
        self.watch_folder = watch_folder  # 目标文件夹名（如NeedApprove）
        
    def on_created(self, event):
        """文件创建事件"""
        if not event.is_directory:
            file_name = os.path.basename(event.src_path)
            if file_name == self.watch_folder:
                self.approver.logger.info(f"检测到新的mbox文件: {event.src_path}")
                # 等待文件写入完成
                time.sleep(1)
                self.approver.process_mbox_file(event.src_path, show_daily_summary=False)
    
    def on_modified(self, event):
        """文件修改事件"""
        if not event.is_directory:
            file_name = os.path.basename(event.src_path)
            if file_name == self.watch_folder:
                self.approver.logger.info(f"检测到mbox文件修改: {event.src_path}")
                time.sleep(1)
                self.approver.process_mbox_file(event.src_path, show_daily_summary=False)

def main():
    """主函数"""
    try:
        # 检查命令行参数
        if len(sys.argv) > 1:
            export_to_file = '--export' in sys.argv or '-e' in sys.argv
            export_excel = '--excel' in sys.argv or '--xlsx' in sys.argv
            export_format = 'xlsx' if export_excel else 'txt'
            
            if sys.argv[1] == '--summary' or sys.argv[1] == '-s':
                # 显示处理汇总报告
                approver = EmailAutoApprover()
                approver.print_processing_summary(export_to_file=export_to_file, export_format=export_format)
                return 0
            elif sys.argv[1] == '--once' or sys.argv[1] == '-o':
                # 一次性处理模式
                approver = EmailAutoApprover()
                success = approver.run_once_mode(export_to_file=export_to_file, export_format=export_format)
                return 0 if success else 1
            elif sys.argv[1] == '--today' or sys.argv[1] == '-t':
                # 显示今日处理汇总报告
                approver = EmailAutoApprover()
                approver.print_daily_processing_summary(export_to_file=export_to_file, export_format=export_format)
                return 0
            elif sys.argv[1] == '--help' or sys.argv[1] == '-h':
                print("ServiceNow邮件自动批准程序")
                print("用法:")
                print("  python email_auto_approve.py         # 启动邮件监控（持续运行）")
                print("  python email_auto_approve.py -o      # 一次性处理模式（处理后退出）")
                print("  python email_auto_approve.py --once  # 一次性处理模式（处理后退出）")
                print("  python email_auto_approve.py -s      # 显示全部处理汇总报告")
                print("  python email_auto_approve.py --summary # 显示全部处理汇总报告")
                print("  python email_auto_approve.py -t      # 显示今日处理汇总报告")
                print("  python email_auto_approve.py --today # 显示今日处理汇总报告")
                print("  python email_auto_approve.py -h      # 显示帮助信息")
                print()
                print("导出选项:")
                print("  -e, --export                          # 导出汇总报告到文件")
                print("  --excel, --xlsx                       # 导出为Excel格式（需要 -e）")
                print()
                print("示例:")
                print("  python email_auto_approve.py -s -e           # 显示并导出全部汇总为txt")
                print("  python email_auto_approve.py -s -e --excel   # 显示并导出全部汇总为Excel")
                print("  python email_auto_approve.py -t -e --xlsx    # 显示并导出今日汇总为Excel")
                print("  python email_auto_approve.py -o -e --excel   # 一次性处理并导出为Excel")
                print()
                print("注意: 导出Excel格式需要安装openpyxl库:")
                print("  pip install openpyxl")
                return 0
        
        approver = EmailAutoApprover()
        
        # 获取监控路径
        watch_folder = approver.config.get('DEFAULT', 'watch_folder')
        
        # 查找Thunderbird邮件文件夹
        full_watch_path = approver.find_thunderbird_mail_folder(watch_folder)
        
        if not full_watch_path:
            print(f"错误: 找不到Thunderbird邮件文件夹: {watch_folder}")
            print("请检查以下内容:")
            print("1. Thunderbird是否已安装并配置")
            print("2. 是否已在Thunderbird中创建文件夹: Archive > ServiceNow > NeedApprove")
            print("3. 配置文件中的thunderbird_profile_path是否正确")
            print(f"4. 当前配置的Thunderbird路径: {approver.config.get('DEFAULT', 'thunderbird_profile_path')}")
            
            # 显示可能的Thunderbird文件夹结构
            profile_path = approver.config.get('DEFAULT', 'thunderbird_profile_path')
            if os.path.exists(profile_path):
                print("\\n发现的Thunderbird配置文件:")
                for item in os.listdir(profile_path):
                    profile_dir = os.path.join(profile_path, item)
                    if os.path.isdir(profile_dir):
                        print(f"  - {item}")
                        mail_dir = os.path.join(profile_dir, 'Mail')
                        if os.path.exists(mail_dir):
                            print(f"    Mail目录存在")
                            local_folders = os.path.join(mail_dir, 'Local Folders')
                            if os.path.exists(local_folders):
                                print(f"    Local Folders目录存在")
            return
        
        print(f"开始监控邮件目录: {full_watch_path}")
        approver.logger.info(f"开始监控邮件目录: {full_watch_path}")
        
        # 获取目标文件夹名
        folder_name = watch_folder.split('/')[-1]  # 获取最后一级文件夹名
        
        # 处理现有邮件
        print("处理现有邮件...")
        mbox_file = os.path.join(full_watch_path, folder_name)
        if os.path.exists(mbox_file) and os.path.isfile(mbox_file):
            approver.process_mbox_file(mbox_file, show_daily_summary=False)
        else:
            print(f"邮件文件不存在: {mbox_file}")
        
        # 启动文件监控
        event_handler = EmailWatcher(approver, folder_name)
        observer = Observer()
        observer.schedule(event_handler, full_watch_path, recursive=False)
        observer.start()
        
        print("邮件监控已启动。按 Ctrl+C 停止...")
        
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            observer.stop()
            print("\n停止监控...")
            
        observer.join()
        
    except Exception as e:
        print(f"程序运行错误: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())